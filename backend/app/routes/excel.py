"""
Excel export routes
"""
from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from typing import Optional
from app.models.database import User, Company, get_db
from app.utils.auth import get_current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io

router = APIRouter(prefix="/api/excel", tags=["excel"])


@router.get("/export")
async def export_companies(
    sehir_filtre: Optional[str] = None,
    ilce_filtre: Optional[str] = None,
    asama_filtre: Optional[str] = None,
    telefon_filtre: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Firmaları Excel'e aktar (ücretsiz)"""
    # Filtrelerle firmaları getir
    query = db.query(Company).filter(Company.user_id == current_user.id)
    
    if sehir_filtre and sehir_filtre != "Hepsi":
        query = query.filter(Company.sehir == sehir_filtre)
    
    if ilce_filtre and ilce_filtre != "Hepsi":
        query = query.filter(Company.ilce == ilce_filtre)
    
    if asama_filtre and asama_filtre != "Hepsi":
        query = query.filter(Company.asama == asama_filtre)
    
    if telefon_filtre == "var":
        query = query.filter(Company.telefon.isnot(None), Company.telefon != "")
    elif telefon_filtre == "yok":
        query = query.filter(
            (Company.telefon.is_(None)) | (Company.telefon == "")
        )
    
    companies = query.order_by(Company.created_at.desc()).all()
    
    if not companies:
        return Response(
            content="Aktarılacak veri bulunamadı",
            status_code=404
        )
    
    # Excel workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Firmalar"
    
    # Başlık satırı
    basliklar = [
        'Sıra', 'Firma Adı', 'Şehir', 'İlçe', 'Kategori', 'Adres', 'Aşama',
        'Telefon', 'Web', 'Rating', 'Değerlendirme Sayısı', 'Fiyat Seviyesi'
    ]
    ws.append(basliklar)
    
    # Başlık stilini ayarla
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Verileri ekle
    for idx, company in enumerate(companies, 1):
        # Fiyat seviyesi formatla
        price_level = company.price_level
        price_str = ''
        if price_level is not None:
            price_str = '₺' * (price_level + 1) if price_level >= 0 else ''
        
        satir = [
            idx,
            company.firma_adi or '',
            company.sehir or '',
            company.ilce or '',
            company.kategori or '',
            company.adres or '',
            company.asama or '',
            company.telefon or '',
            company.web or '',
            company.rating or '',
            company.user_ratings_total or '',
            price_str
        ]
        ws.append(satir)
    
    # Kolon genişliklerini ayarla
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    
    # Satır yüksekliğini ayarla
    ws.row_dimensions[1].height = 25
    
    # Excel dosyasını memory'ye yaz
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Dosya adı
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"firmalar_{timestamp}.xlsx"
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

